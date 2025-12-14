# app/excel.py
import datetime as dt
from pathlib import Path
from openpyxl import load_workbook

LOG_SHEET_INDEX = 0  # первый лист
START_DAY_COL = 4   # колонка D = день 1
MAX_HOURS = 8

def _ceil_hours(minutes: int) -> int:
    if minutes <= 0:
        return 0
    return min((minutes + 59) // 60, MAX_HOURS)

def append_log(
    wb_path: Path,
    ts: dt.datetime,
    telegram_id: int,
    fio: str,
    action: str,
    geo_url: str,
    object_name: str | None
):
    wb = load_workbook(wb_path)
    ws = wb.worksheets[LOG_SHEET_INDEX]

    ws.append([
        ts.strftime("%Y-%m-%d %H:%M:%S"),
        telegram_id,
        fio,
        action,
        geo_url,
        object_name or ""
    ])

    wb.save(wb_path)

def write_object_day(
    wb_path: Path,
    object_name: str,
    fio: str,
    day: int,
    value: int | str
):
    wb = load_workbook(wb_path)
    ws = wb[object_name]

    fio_row = None
    for r in range(3, ws.max_row + 1):
        if str(ws.cell(r, 2).value).strip() == fio.strip():
            fio_row = r
            break

    if fio_row is None:
        raise ValueError(f"ФИО не найдено в объекте {object_name}")

    col = START_DAY_COL + day - 1
    ws.cell(fio_row, col).value = value

    _recalc_totals(ws, fio_row)

    wb.save(wb_path)

def _recalc_totals(ws, row: int):
    total_hours = 0
    total_days = 0

    for c in range(START_DAY_COL, ws.max_column):
        v = ws.cell(row, c).value
        if isinstance(v, (int, float)):
            total_hours += int(v)
            if int(v) > 0:
                total_days += 1
        elif isinstance(v, str) and v.upper() == "Б":
            total_days += 1

    ws.cell(row, ws.max_column - 1).value = total_hours
    ws.cell(row, ws.max_column).value = total_days
